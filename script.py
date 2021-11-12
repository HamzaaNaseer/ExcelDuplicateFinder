import openpyxl as xl
wb = input('enter the file name : ')
wb = xl.load_workbook(wb + '.xlsx');

print('file loaded sucessfully')


sheet = input('enter the sheet name : ')
#'Total Labour'
sheet = wb[sheet]

startingRow = int(input('enter the row number from where you want to start the comparison: '))
totalRows = int(input('enter total rows: '))
column = int(input('enter column number whose valules will be compared ,(first column has index 0): ' ))


#column 8


#3,4023
for i in range (startingRow,totalRows):
    cell = sheet.cell(i,column)
    data = str(cell.value)
    dataLower = data.lower()
    
    #print(dataLower)
    for j in range (3,4023):
        iterCell = sheet.cell(j,column)
        iterData = str(iterCell.value)
        iterLower = iterData.lower()
        
        
        if(dataLower == iterLower):
            outerDuplicate = sheet.cell(i,1)
            outerDuplicate = outerDuplicate.value
            duplicateData = sheet.cell(j,1)
            duplicateData = duplicateData.value
            #print('dataLower',dataLower,'iterLower',iterLower)
            if(outerDuplicate == duplicateData or dataLower == 'none'):
                continue
            
            print('duplicate data fount at',duplicateData,"----",outerDuplicate)


print('done')
        

